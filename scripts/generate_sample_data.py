"""
Generate sample startup data for testing MinDD.

Creates:
  data/startups/alpha/  - AlphaFlow (B2B SaaS)
    pitch_deck.pdf
    investor_update.pdf
    financial_model.xlsx

  data/startups/beta/   - BetaMart (E-commerce marketplace)
    pitch_deck.pdf
    investor_update.pdf
    financial_model.xlsx

Run:
    python scripts/generate_sample_data.py
"""

from __future__ import annotations

import sys
from pathlib import Path

# Make sure project root is on path
sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fpdf import FPDF


# ---------------------------------------------------------------------------
# PDF helpers
# ---------------------------------------------------------------------------

def _ascii(text: str) -> str:
    """Replace non-latin-1 chars with ASCII equivalents."""
    return (
        text.replace("\u2013", "-")
            .replace("\u2014", "--")
            .replace("\u2018", "'")
            .replace("\u2019", "'")
            .replace("\u201c", '"')
            .replace("\u201d", '"')
            .replace("\u2022", "*")
            .encode("latin-1", errors="replace")
            .decode("latin-1")
    )


class StartupPDF(FPDF):
    def __init__(self, startup_name: str):
        super().__init__()
        self.startup_name = _ascii(startup_name)

    def header(self):
        self.set_font("Helvetica", "B", 14)
        self.cell(0, 10, self.startup_name, new_x="LMARGIN", new_y="NEXT", align="C")
        self.ln(2)

    def add_section(self, title: str, body: str):
        self.set_font("Helvetica", "B", 11)
        self.cell(0, 8, _ascii(title), new_x="LMARGIN", new_y="NEXT")
        self.set_font("Helvetica", "", 10)
        self.multi_cell(0, 6, _ascii(body))
        self.ln(3)


def generate_pdf(path: Path, startup_name: str, sections: list[tuple[str, str]]):
    pdf = StartupPDF(startup_name)
    pdf.add_page()
    for title, body in sections:
        pdf.add_section(title, body)
    pdf.output(str(path))
    print(f"  Created: {path}")


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
METRIC_FILL = PatternFill("solid", fgColor="E2EFDA")


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")


def style_subheader(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = Font(bold=True)


# ---------------------------------------------------------------------------
# Alpha – AlphaFlow (B2B SaaS)
# ---------------------------------------------------------------------------

ALPHA_DIR = Path("data/startups/alpha")

ALPHA_PITCH_SECTIONS = [
    ("Company Overview",
     "AlphaFlow is a B2B SaaS workflow automation platform that helps mid-market companies "
     "eliminate manual processes and reduce operational overhead. Founded in 2022, AlphaFlow "
     "serves over 120 enterprise customers across financial services, healthcare, and logistics."),

    ("Problem",
     "Enterprise teams waste 30-40% of their time on repetitive manual workflows. Existing "
     "solutions are either too rigid (legacy BPM tools) or too lightweight (no-code tools that "
     "break at scale). There is no modern, API-first workflow platform for mid-market companies."),

    ("Solution",
     "AlphaFlow provides a drag-and-drop workflow builder with deep integrations (Salesforce, "
     "HubSpot, SAP), an AI-powered anomaly detection layer, and enterprise-grade audit trails. "
     "Customers go live in under 2 weeks and report average efficiency gains of 45%."),

    ("Traction",
     "ARR: $960K as of Q4 2025, growing 60% year-over-year. NRR: 118%. "
     "Paying customers: 123. Average contract value: $7,800/year. "
     "Churn rate: 4.2% annually. CAC: $3,200. LTV: $22,500."),

    ("Market Opportunity",
     "The global BPM and workflow automation market is projected to reach $26B by 2027 (CAGR 12%). "
     "AlphaFlow targets the underserved mid-market segment (500-5,000 employees), representing "
     "an addressable market of $4.2B."),

    ("Business Model",
     "SaaS subscription with per-seat pricing: $65/user/month (Professional), $95/user/month (Enterprise). "
     "Implementation and onboarding services contribute 15% of revenue. "
     "Annual contracts with monthly billing option."),

    ("Competition",
     "Primary competitors: Zapier (SMB-focused, limited enterprise features), "
     "ServiceNow (complex, expensive, 6-12 month implementation), "
     "UiPath (RPA-focused, not workflow-native). "
     "AlphaFlow's key differentiators: 2-week deployment, mid-market pricing, native AI layer."),

    ("Key Risks",
     "1. Enterprise sales cycles can extend to 6+ months, impacting cash flow predictability. "
     "2. Dependence on third-party API integrations creates vendor risk. "
     "3. Competition from well-funded players like ServiceNow entering the mid-market. "
     "4. Talent acquisition in competitive ML/AI engineering market. "
     "5. Data security and compliance requirements vary by industry vertical."),

    ("Team",
     "CEO: Sarah Chen – ex-Salesforce, 10 years enterprise software. "
     "CTO: James Park – ex-Palantir, distributed systems expert. "
     "VP Sales: Maria Rodriguez – ex-HubSpot, built 0-to-$10M ARR at two prior startups. "
     "Total headcount: 28 (18 engineering, 6 sales, 4 operations)."),

    ("Funding Ask",
     "Raising $3.5M Series A to fund 18-month runway. "
     "Use of funds: 55% engineering (AI/ML features, platform scaling), "
     "30% sales & marketing (2 new AEs, demand gen), "
     "15% operations (compliance, security certifications ISO 27001, SOC 2)."),
]

ALPHA_UPDATE_SECTIONS = [
    ("Q4 2025 Investor Update – AlphaFlow",
     "Dear Investors, We are pleased to share AlphaFlow's Q4 2025 performance. "
     "This was our strongest quarter to date."),

    ("Financial Highlights",
     "MRR: $80,000 (up from $65,000 in Q3, +23% QoQ). ARR: $960,000. "
     "Monthly burn rate: $92,000. Cash balance: $1.15M. Runway: 12.5 months. "
     "Gross margin: 74%. Net Revenue Retention: 118%."),

    ("Key Wins",
     "Signed 3 new enterprise contracts: LogiCorp ($18K ACV), FinServ Holdings ($22K ACV), "
     "MedPath Systems ($15K ACV). Pipeline is at $1.2M ARR across 45 qualified opportunities. "
     "Closed Q4 with 123 paying customers vs 98 in Q3."),

    ("Product Updates",
     "Launched AI anomaly detection (v2.1) with 94% precision. "
     "Released Salesforce bidirectional sync – reduced integration time from 3 days to 2 hours. "
     "SOC 2 Type II certification completed in November 2025."),

    ("Challenges",
     "Enterprise sales cycle averaging 4.5 months (target: 3 months). "
     "Engineering hiring behind plan: 3 of 5 planned Q4 hires completed. "
     "One churn event: $14K ACV customer acquired by larger company using ServiceNow."),

    ("Q1 2026 Outlook",
     "MRR target: $95,000. Pipeline conversion target: 8 new enterprise logos. "
     "Engineering: complete real-time collaboration feature (high-demand customer request). "
     "Finance: targeting $3.5M Series A close by end of Q1 2026."),
]


def create_alpha_excel(path: Path):
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Revenue Model ----
    ws = wb.active
    ws.title = "Revenue Model"

    months = ["Jan-25", "Feb-25", "Mar-25", "Apr-25", "May-25", "Jun-25",
              "Jul-25", "Aug-25", "Sep-25", "Oct-25", "Nov-25", "Dec-25",
              "Jan-26", "Feb-26", "Mar-26", "Apr-26", "May-26", "Jun-26"]

    headers = ["Metric"] + months
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col))

    mrr_values = [50000, 52000, 54000, 56000, 58000, 60000,
                  63000, 66000, 70000, 74000, 77000, 80000,
                  84000, 88000, 92000, 96000, 100000, 105000]

    customers = [80, 82, 85, 88, 91, 94, 98, 102, 108, 114, 119, 123,
                 128, 134, 140, 146, 152, 159]

    ws.append(["Monthly Recurring Revenue (MRR)"] + mrr_values)
    ws.append(["Annual Recurring Revenue (ARR)"] + [v * 12 for v in mrr_values])
    ws.append(["Paying Customers"] + customers)
    ws.append(["Average Contract Value (ACV, monthly)"] + [round(m / c, 0) for m, c in zip(mrr_values, customers)])

    # Revenue Growth row using formulas
    growth_row = ["MoM Revenue Growth (%)"]
    for i, month in enumerate(months):
        if i == 0:
            growth_row.append("N/A")
        else:
            col_letter_curr = get_column_letter(i + 2)
            col_letter_prev = get_column_letter(i + 1)
            # Formula: (current - previous) / previous * 100
            growth_row.append(f"=({col_letter_curr}2-{col_letter_prev}2)/{col_letter_prev}2*100")
    ws.append(growth_row)

    ws.column_dimensions["A"].width = 35
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # ---- Sheet 2: Expenses & Burn ----
    ws2 = wb.create_sheet("Expenses & Burn")

    exp_headers = ["Expense Category"] + months
    ws2.append(exp_headers)
    for col, _ in enumerate(exp_headers, 1):
        style_header(ws2.cell(row=1, column=col))

    salary = [55000, 55000, 58000, 58000, 62000, 62000,
              65000, 68000, 70000, 72000, 75000, 78000,
              80000, 82000, 85000, 88000, 90000, 92000]
    cogs   = [13000, 13500, 14000, 14500, 15000, 15500,
              16400, 17200, 18200, 19200, 20000, 20800,
              21800, 22900, 23900, 25000, 26000, 27300]
    infra  = [5000,  5000,  5200,  5200,  5500,  5500,
              5800,  6000,  6200,  6400,  6600,  6800,
              7000,  7200,  7500,  7800,  8000,  8200]
    marketing = [8000, 8000, 9000, 9500, 10000, 10000,
                 10500, 11000, 11500, 12000, 12500, 13000,
                 13500, 14000, 14500, 15000, 15500, 16000]

    ws2.append(["Salaries & Benefits"] + salary)
    ws2.append(["Cost of Goods Sold (COGS)"] + cogs)
    ws2.append(["Infrastructure & Hosting"] + infra)
    ws2.append(["Sales & Marketing"] + marketing)

    # Total expenses formula
    total_row = ["Total Expenses"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        total_row.append(f"={col}2+{col}3+{col}4+{col}5")
    ws2.append(total_row)

    # Burn rate = Expenses - Revenue
    burn_row = ["Net Monthly Burn Rate"]
    mrr_sht = "Revenue Model"
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        burn_row.append(f"={col}6-'{mrr_sht}'!{col}2")
    ws2.append(burn_row)

    # Gross Margin formula
    gm_row = ["Gross Margin (%)"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        gm_row.append(f"=('Revenue Model'!{col}2-{col}3)/'Revenue Model'!{col}2*100")
    ws2.append(gm_row)

    ws2.column_dimensions["A"].width = 35

    # ---- Sheet 3: Cash & Runway ----
    ws3 = wb.create_sheet("Cash & Runway")

    cash_headers = ["Metric"] + months
    ws3.append(cash_headers)
    for col, _ in enumerate(cash_headers, 1):
        style_header(ws3.cell(row=1, column=col))

    opening_cash = [2000000]
    for i in range(1, len(months)):
        opening_cash.append(None)  # will use formulas

    ws3.append(["Opening Cash Balance"] + [opening_cash[0]] + [None] * (len(months) - 1))

    # Ending cash = Opening Cash - Burn
    ending_row = ["Ending Cash Balance"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        if i == 0:
            ending_row.append(f"={col}2-'Expenses & Burn'!{col}7")
        else:
            prev_col = get_column_letter(i + 1)
            ending_row.append(f"={prev_col}3-'Expenses & Burn'!{col}7")
    ws3.append(ending_row)

    # Runway = Ending Cash / Monthly Burn
    runway_row = ["Runway (months)"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        runway_row.append(f"={col}3/'Expenses & Burn'!{col}7")
    ws3.append(runway_row)

    ws3.column_dimensions["A"].width = 35

    # ---- Sheet 4: Key Metrics Summary ----
    ws4 = wb.create_sheet("Key Metrics Summary")
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 15

    summary_data = [
        ("Metric", "Value", "Unit"),
        ("ARR (Q4 2025)", 960000, "USD"),
        ("MRR (Dec 2025)", 80000, "USD"),
        ("Monthly Burn Rate", 92000, "USD"),
        ("Cash Balance", 1150000, "USD"),
        ("Runway", "=B5/B4", "months"),
        ("Gross Margin", 74, "%"),
        ("YoY Revenue Growth", 60, "%"),
        ("Net Revenue Retention (NRR)", 118, "%"),
        ("Customer Acquisition Cost (CAC)", 3200, "USD"),
        ("Lifetime Value (LTV)", 22500, "USD"),
        ("LTV:CAC Ratio", "=B11/B10", "x"),
        ("Paying Customers", 123, "count"),
        ("Annual Churn Rate", 4.2, "%"),
        ("Headcount", 28, "employees"),
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                style_header(cell)
            elif col_idx == 1:
                style_subheader(cell)

    wb.save(str(path))
    print(f"  Created: {path}")


# ---------------------------------------------------------------------------
# Beta – BetaMart (E-commerce marketplace)
# ---------------------------------------------------------------------------

BETA_DIR = Path("data/startups/beta")

BETA_PITCH_SECTIONS = [
    ("Company Overview",
     "BetaMart is an e-commerce marketplace connecting independent artisan brands with "
     "conscious consumers in the US and Canada. Founded in 2021, BetaMart has grown to "
     "host 2,400 active sellers and 85,000 monthly active buyers."),

    ("Problem",
     "Independent artisan brands struggle to compete with mass-market retailers on large "
     "platforms like Amazon and Etsy, where algorithmic disadvantages and high fees (15-25%) "
     "erode margins. Consumers seeking authentic, ethical products lack a curated discovery layer."),

    ("Solution",
     "BetaMart provides a curated, values-aligned marketplace with: lower take rate (8%), "
     "a brand story platform, AI-powered personalized discovery, and sustainability verification. "
     "Sellers report 3x higher conversion rates vs general marketplaces."),

    ("Traction",
     "GMV (Gross Merchandise Value): $4.2M in 2025. Revenue: $336K (8% take rate). "
     "MRR: $28,000. ARR: $336,000. Active sellers: 2,400. MAU buyers: 85,000. "
     "Average order value: $67. Repeat purchase rate: 38%."),

    ("Market Opportunity",
     "The global artisan/handmade e-commerce market is valued at $680B and growing at 9% CAGR. "
     "Ethical shopping is a mega-trend: 73% of Gen Z prefer sustainable brands. "
     "Serviceable addressable market: $12B in North America."),

    ("Business Model",
     "Transaction fee: 8% on GMV. Premium seller subscriptions: $29/month (Pro). "
     "Promoted listings and advertising: 15% of revenue. "
     "Revenue growth driven by GMV growth and seller tier upgrades."),

    ("Competition",
     "Etsy: dominant but commoditized, 6.5% fee + payment processing. "
     "Faire: B2B wholesale, different segment. "
     "Amazon Handmade: poor discovery, 15% fee. "
     "BetaMart's moat: curation quality, lower fees, sustainability verification, community."),

    ("Key Risks",
     "1. Seller concentration: top 20% of sellers represent 65% of GMV – churn risk. "
     "2. Fraud and counterfeit listings require ongoing trust & safety investment. "
     "3. Customer acquisition cost is rising as paid social CPMs increase 30% YoY. "
     "4. Seasonality: 45% of GMV in Q4, creating cash flow variability. "
     "5. Etsy or Amazon could copy curation features."),

    ("Team",
     "CEO: Alex Thompson – ex-Shopify, built marketplace product 0-to-$100M GMV. "
     "CPO: Priya Mehta – ex-Pinterest, discovery and recommendation systems. "
     "Head of Operations: David Kim – ex-Faire, seller success and logistics. "
     "Total headcount: 19 (8 engineering, 5 operations, 3 marketing, 3 management)."),

    ("Funding Ask",
     "Raising $2M Seed extension to reach 18-month runway and GMV/revenue breakeven. "
     "Use of funds: 40% product & engineering (mobile app, AI recommendations), "
     "35% seller acquisition (onboarding team, content creation grants), "
     "25% marketing (brand awareness, influencer partnerships)."),
]

BETA_UPDATE_SECTIONS = [
    ("Q4 2025 Investor Update – BetaMart",
     "Team, We are pleased to share a strong close to 2025. Q4 was our highest-GMV "
     "quarter ever at $1.8M, driven by holiday shopping and our curated gift guide campaign."),

    ("Financial Highlights",
     "MRR: $28,000 (transaction fees + subscriptions). ARR: $336,000. "
     "Q4 GMV: $1.8M. Full-year GMV 2025: $4.2M (vs $2.1M in 2024, +100% YoY). "
     "Monthly burn rate: $58,000. Cash balance: $820,000. Runway: 14.1 months. "
     "Gross margin: 62%. Take rate: 8.0%."),

    ("Key Wins",
     "Holiday gift guide drove $420K GMV in December alone. "
     "Launched Premium Seller tier ($29/month): 340 subscribers in 60 days. "
     "Partnered with 3 ethical fashion influencers (combined 2.1M followers). "
     "App download: 12,000 new iOS downloads in Q4 from App Store feature."),

    ("Product Updates",
     "Launched AI-powered 'Discover for You' feed – 28% increase in session depth. "
     "Sustainability Score badge rolled out to 1,800 verified sellers. "
     "Mobile app v2.0: checkout conversion improved by 18%."),

    ("Challenges",
     "CAC increased to $4.20 per buyer (target: $3.50) due to rising Meta CPMs. "
     "Seller churn Q4: 85 sellers (3.5% of active base) – mostly due to inventory issues. "
     "Logistics partner delays in November caused 120 late deliveries, NPS impact."),

    ("2026 Outlook",
     "GMV target 2026: $7.5M (+79% YoY). Revenue target: $600K. "
     "Seller base target: 4,000. MAU target: 150,000 buyers. "
     "Planning Series A raise of $4M in Q2 2026 upon hitting $500K ARR milestone."),
]


def create_beta_excel(path: Path):
    wb = openpyxl.Workbook()

    months = ["Jan-25", "Feb-25", "Mar-25", "Apr-25", "May-25", "Jun-25",
              "Jul-25", "Aug-25", "Sep-25", "Oct-25", "Nov-25", "Dec-25",
              "Jan-26", "Feb-26", "Mar-26", "Apr-26", "May-26", "Jun-26"]

    # ---- Sheet 1: GMV & Revenue ----
    ws = wb.active
    ws.title = "GMV & Revenue"

    gmv_values = [180000, 190000, 200000, 230000, 250000, 270000,
                  290000, 310000, 340000, 380000, 430000, 500000,  # holiday spike Dec
                  520000, 540000, 570000, 600000, 630000, 660000]

    take_rate = 0.08
    revenue_values = [round(g * take_rate) for g in gmv_values]
    seller_counts = [1800, 1850, 1900, 1980, 2050, 2120,
                     2200, 2280, 2350, 2380, 2410, 2400,
                     2450, 2510, 2580, 2660, 2740, 2820]

    headers = ["Metric"] + months
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col))

    ws.append(["Gross Merchandise Value (GMV)"] + gmv_values)
    ws.append(["Platform Revenue (8% take rate)"] + revenue_values)
    ws.append(["Premium Seller Subscriptions ($29/mo)"] +
              [0]*12 + [9860, 10150, 10440, 10730, 11020, 11310])
    ws.append(["Total Revenue"] + [None] * len(months))
    # Total Revenue formula
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        ws.cell(row=5, column=i + 2, value=f"={col}3+{col}4")

    ws.append(["Active Sellers"] + seller_counts)

    # MoM GMV Growth
    growth_row = ["MoM GMV Growth (%)"]
    for i in range(len(months)):
        if i == 0:
            growth_row.append("N/A")
        else:
            col_curr = get_column_letter(i + 2)
            col_prev = get_column_letter(i + 1)
            growth_row.append(f"=({col_curr}2-{col_prev}2)/{col_prev}2*100")
    ws.append(growth_row)

    ws.column_dimensions["A"].width = 40
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # ---- Sheet 2: Expenses ----
    ws2 = wb.create_sheet("Expenses & Burn")

    exp_headers = ["Expense Category"] + months
    ws2.append(exp_headers)
    for col, _ in enumerate(exp_headers, 1):
        style_header(ws2.cell(row=1, column=col))

    salary =     [28000]*6 + [30000]*6 + [32000]*6
    tech_infra = [5000]*6  + [5500]*6  + [6000]*6
    marketing =  [12000, 12000, 13000, 14000, 15000, 15000,
                  16000, 17000, 18000, 19000, 20000, 22000,
                  20000, 20000, 21000, 22000, 22000, 23000]
    ops =        [5000]*18

    ws2.append(["Salaries & Benefits"] + salary)
    ws2.append(["Technology & Infrastructure"] + tech_infra)
    ws2.append(["Sales & Marketing"] + marketing)
    ws2.append(["Operations & Support"] + ops)

    total_exp_row = ["Total Expenses"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        total_exp_row.append(f"={col}2+{col}3+{col}4+{col}5")
    ws2.append(total_exp_row)

    burn_row = ["Net Monthly Burn Rate"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        burn_row.append(f"={col}6-'GMV & Revenue'!{col}5")
    ws2.append(burn_row)

    gm_row = ["Gross Margin (%)"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        gm_row.append(f"=('GMV & Revenue'!{col}5-{col}3)/'GMV & Revenue'!{col}5*100")
    ws2.append(gm_row)

    ws2.column_dimensions["A"].width = 35

    # ---- Sheet 3: Cash & Runway ----
    ws3 = wb.create_sheet("Cash & Runway")

    cash_headers = ["Metric"] + months
    ws3.append(cash_headers)
    for col, _ in enumerate(cash_headers, 1):
        style_header(ws3.cell(row=1, column=col))

    ws3.append(["Opening Cash Balance", 1500000] + [None] * (len(months) - 1))

    ending_row = ["Ending Cash Balance"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        if i == 0:
            ending_row.append(f"={col}2-'Expenses & Burn'!{col}7")
        else:
            prev_col = get_column_letter(i + 1)
            ending_row.append(f"={prev_col}3-'Expenses & Burn'!{col}7")
    ws3.append(ending_row)

    runway_row = ["Runway (months)"]
    for i in range(len(months)):
        col = get_column_letter(i + 2)
        runway_row.append(f"={col}3/'Expenses & Burn'!{col}7")
    ws3.append(runway_row)

    ws3.column_dimensions["A"].width = 35

    # ---- Sheet 4: Key Metrics Summary ----
    ws4 = wb.create_sheet("Key Metrics Summary")
    ws4.column_dimensions["A"].width = 40
    ws4.column_dimensions["B"].width = 20
    ws4.column_dimensions["C"].width = 15

    summary_data = [
        ("Metric", "Value", "Unit"),
        ("ARR (Q4 2025)", 336000, "USD"),
        ("MRR (Dec 2025)", 28000, "USD"),
        ("GMV Full Year 2025", 4200000, "USD"),
        ("Monthly Burn Rate", 58000, "USD"),
        ("Cash Balance", 820000, "USD"),
        ("Runway", "=B6/B5", "months"),
        ("Gross Margin", 62, "%"),
        ("YoY GMV Growth", 100, "%"),
        ("YoY Revenue Growth", 108, "%"),
        ("Take Rate", 8, "%"),
        ("Active Sellers", 2400, "count"),
        ("Monthly Active Buyers (MAU)", 85000, "count"),
        ("Customer Acquisition Cost (CAC)", 4.20, "USD"),
        ("Average Order Value (AOV)", 67, "USD"),
        ("Repeat Purchase Rate", 38, "%"),
        ("Annual Churn Rate (Sellers)", 12, "%"),
        ("Headcount", 19, "employees"),
    ]

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                style_header(cell)
            elif col_idx == 1:
                style_subheader(cell)

    wb.save(str(path))
    print(f"  Created: {path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Generating sample startup data...\n")

    # Alpha
    ALPHA_DIR.mkdir(parents=True, exist_ok=True)
    print("AlphaFlow (B2B SaaS):")
    generate_pdf(ALPHA_DIR / "pitch_deck.pdf", "AlphaFlow - Pitch Deck", ALPHA_PITCH_SECTIONS)
    generate_pdf(ALPHA_DIR / "investor_update.pdf", "AlphaFlow - Q4 2025 Investor Update", ALPHA_UPDATE_SECTIONS)
    create_alpha_excel(ALPHA_DIR / "financial_model.xlsx")

    print()

    # Beta
    BETA_DIR.mkdir(parents=True, exist_ok=True)
    print("BetaMart (E-commerce Marketplace):")
    generate_pdf(BETA_DIR / "pitch_deck.pdf", "BetaMart - Pitch Deck", BETA_PITCH_SECTIONS)
    generate_pdf(BETA_DIR / "investor_update.pdf", "BetaMart - Q4 2025 Investor Update", BETA_UPDATE_SECTIONS)
    create_beta_excel(BETA_DIR / "financial_model.xlsx")

    print("\nSample data generation complete.")
    print("Next step: POST /ingest for each startup, or run scripts/run_evaluation.py")


if __name__ == "__main__":
    main()
