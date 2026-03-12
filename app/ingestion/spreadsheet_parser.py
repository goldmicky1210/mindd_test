"""
Parse Excel spreadsheets: extract values, formulas, and derive structured
financial metrics.  Returns both a list of text chunks (for embedding) and a
dict of structured metrics (for the metadata store).
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_spreadsheet(file_path: str | Path) -> dict[str, Any]:
    """
    Parse an Excel workbook and return:
      - "text_chunks"  : list[str]  – human-readable descriptions for embedding
      - "metrics"      : list[dict] – structured financial metrics
      - "sheets"       : list[dict] – raw sheet data (values + formulas)
    """
    path = Path(file_path)
    wb_values = openpyxl.load_workbook(path, data_only=True)
    wb_formulas = openpyxl.load_workbook(path, data_only=False)

    sheets_data = []
    all_metrics: list[dict] = []
    text_chunks: list[str] = []

    for sheet_name in wb_values.sheetnames:
        ws_values = wb_values[sheet_name]
        ws_formulas = wb_formulas[sheet_name]

        sheet_info = _extract_sheet(ws_values, ws_formulas, sheet_name)
        sheets_data.append(sheet_info)

        metrics = _extract_metrics(sheet_info, sheet_name, str(path.name))
        all_metrics.extend(metrics)

        chunk = _sheet_to_text(sheet_info, sheet_name)
        if chunk:
            text_chunks.append(chunk)

    wb_values.close()
    wb_formulas.close()

    # Add a summary chunk with all extracted metrics
    if all_metrics:
        text_chunks.append(_metrics_summary(all_metrics, path.name))

    return {
        "text_chunks": text_chunks,
        "metrics": all_metrics,
        "sheets": sheets_data,
    }


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _extract_sheet(ws_values, ws_formulas, sheet_name: str) -> dict:
    """Build a cell table with values and formulas."""
    rows = []
    headers: list[str] = []

    max_row = ws_values.max_row or 0
    max_col = ws_values.max_column or 0

    if max_row == 0 or max_col == 0:
        return {"name": sheet_name, "headers": [], "rows": [], "cells": {}}

    # Attempt to identify header row (first non-empty row)
    for row_idx, row in enumerate(ws_values.iter_rows(min_row=1, max_row=min(3, max_row)), start=1):
        values = [str(c.value).strip() if c.value is not None else "" for c in row]
        if any(values):
            headers = values
            data_start = row_idx + 1
            break
    else:
        data_start = 2

    # Build cell dict: coord -> {value, formula}
    cells: dict[str, dict] = {}
    for row_v, row_f in zip(
        ws_values.iter_rows(min_row=1, max_row=max_row),
        ws_formulas.iter_rows(min_row=1, max_row=max_row),
    ):
        row_data: list[dict] = []
        for cv, cf in zip(row_v, row_f):
            coord = cv.coordinate
            value = cv.value
            formula = str(cf.value) if str(cf.value or "").startswith("=") else None
            cells[coord] = {"value": value, "formula": formula}
            row_data.append({"coord": coord, "value": value, "formula": formula})
        rows.append(row_data)

    return {
        "name": sheet_name,
        "headers": headers,
        "rows": rows,
        "cells": cells,
        "data_start": data_start,
    }


_METRIC_KEYWORDS: dict[str, list[str]] = {
    "arr": ["arr", "annual recurring revenue", "annualized revenue"],
    "mrr": ["mrr", "monthly recurring revenue"],
    "burn_rate": ["burn rate", "monthly burn", "net burn", "cash burn"],
    "runway": ["runway", "months of runway", "cash runway"],
    "revenue": ["total revenue", "revenue", "net revenue", "gross revenue"],
    "revenue_growth": ["revenue growth", "growth rate", "yoy growth", "mom growth"],
    "gross_margin": ["gross margin", "gm %", "gross profit margin"],
    "cac": ["cac", "customer acquisition cost"],
    "ltv": ["ltv", "lifetime value", "customer ltv"],
    "churn": ["churn", "churn rate", "monthly churn"],
    "cash": ["cash", "cash balance", "cash on hand", "ending cash"],
    "expenses": ["total expenses", "opex", "operating expenses", "expenses"],
    "ebitda": ["ebitda", "operating income"],
    "headcount": ["headcount", "employees", "team size", "fte"],
}


def _extract_metrics(sheet_info: dict, sheet_name: str, source_file: str) -> list[dict]:
    """Scan rows for recognised financial metric names and extract their values."""
    metrics: list[dict] = []
    seen_keys: set[str] = set()

    for row in sheet_info.get("rows", []):
        if not row:
            continue

        # Look for a label cell (typically column A or B)
        for cell in row[:3]:
            cell_text = str(cell.get("value") or "").lower().strip()
            if not cell_text:
                continue

            for metric_key, keywords in _METRIC_KEYWORDS.items():
                if any(kw in cell_text for kw in keywords):
                    # Grab numeric value from next cell(s)
                    numeric_value = None
                    numeric_text = None
                    period = None

                    for other_cell in row[1:]:
                        v = other_cell.get("value")
                        if v is not None and isinstance(v, (int, float)):
                            numeric_value = float(v)
                            numeric_text = _format_value(v, metric_key)
                            break
                        elif v is not None and isinstance(v, str):
                            # Try to extract number from string like "$1.2M"
                            cleaned = re.sub(r"[^0-9.\-]", "", v)
                            if cleaned:
                                try:
                                    numeric_value = float(cleaned)
                                    numeric_text = v
                                    break
                                except ValueError:
                                    pass

                    # Avoid duplicate metrics per sheet
                    key = f"{sheet_name}::{metric_key}"
                    if key not in seen_keys and (numeric_value is not None or numeric_text):
                        seen_keys.add(key)
                        metrics.append(
                            {
                                "metric_name": metric_key,
                                "display_name": cell["value"],
                                "value": numeric_value,
                                "value_text": numeric_text,
                                "unit": _infer_unit(metric_key),
                                "period": period,
                                "source_file": source_file,
                                "sheet": sheet_name,
                            }
                        )
                    break  # matched keyword – stop checking other keywords for this label cell

    return metrics


def _sheet_to_text(sheet_info: dict, sheet_name: str) -> str:
    """Convert a sheet into a human-readable text block for embedding."""
    lines = [f"Sheet: {sheet_name}"]

    if sheet_info.get("headers"):
        lines.append("Columns: " + " | ".join(h for h in sheet_info["headers"] if h))

    for row in sheet_info.get("rows", [])[:60]:  # cap at 60 rows
        parts = []
        for cell in row:
            v = cell.get("value")
            f = cell.get("formula")
            if v is not None:
                parts.append(str(v))
            elif f:
                parts.append(f"[formula: {f}]")
        row_text = " | ".join(p for p in parts if p)
        if row_text:
            lines.append(row_text)

    return "\n".join(lines)


def _metrics_summary(metrics: list[dict], filename: str) -> str:
    """Generate a concise metrics summary paragraph."""
    lines = [f"Financial metrics extracted from {filename}:"]
    for m in metrics:
        display = m.get("display_name") or m["metric_name"].replace("_", " ").title()
        val = m.get("value_text") or (f"{m['value']}" if m.get("value") is not None else "N/A")
        unit = m.get("unit") or ""
        period = f" ({m['period']})" if m.get("period") else ""
        lines.append(f"  - {display}: {val} {unit}{period}".strip())
    return "\n".join(lines)


def _format_value(v: float, metric_key: str) -> str:
    if metric_key in ("gross_margin", "revenue_growth", "churn"):
        return f"{v:.1f}%"
    if metric_key == "runway":
        return f"{v:.1f} months"
    if v >= 1_000_000:
        return f"${v/1_000_000:.2f}M"
    if v >= 1_000:
        return f"${v/1_000:.1f}K"
    return str(v)


def _infer_unit(metric_key: str) -> str:
    pct_metrics = {"gross_margin", "revenue_growth", "churn"}
    month_metrics = {"runway"}
    if metric_key in pct_metrics:
        return "%"
    if metric_key in month_metrics:
        return "months"
    return "USD"
